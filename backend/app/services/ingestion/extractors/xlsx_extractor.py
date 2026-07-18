from __future__ import annotations

from collections.abc import Callable
from io import BytesIO
from typing import Any, cast

from app.core.errors import ApiError
from app.services.ingestion.extractors.base import (
    BaseExtractor,
    ExtractionRequest,
    ExtractionResult,
)
from app.services.ingestion.parser_service import sanitize_document_text


class XlsxExtractor(BaseExtractor):
    extraction_method = "xlsx_native"
    supported_extensions = frozenset({".xlsx"})
    supported_mime_types = frozenset(
        {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
    )

    def __init__(self, *, max_text_chars: int) -> None:
        self.max_text_chars = max_text_chars

    def extract(self, request: ExtractionRequest) -> ExtractionResult:
        loader = self._load_workbook_loader()
        try:
            workbook = cast(
                Any,
                loader(
                    filename=BytesIO(request.payload), read_only=True, data_only=True
                ),
            )
        except Exception as exc:  # noqa: BLE001
            raise ApiError(
                code="XLSX_PARSE_FAILED",
                message="Failed to parse XLSX workbook.",
                status_code=422,
            ) from exc

        lines: list[str] = []
        sheet_count = 0
        for worksheet in workbook.worksheets:
            sheet_count += 1
            lines.append(f"## Sheet: {worksheet.title}")
            for row_idx, row in enumerate(
                worksheet.iter_rows(values_only=True), start=1
            ):
                values: list[str] = []
                for col_idx, value in enumerate(row, start=1):
                    if value is None:
                        continue
                    normalized = sanitize_document_text(str(value)).strip()
                    if not normalized:
                        continue
                    values.append(f"C{col_idx}={normalized}")
                if values:
                    lines.append(f"R{row_idx}: " + " | ".join(values))

        text = "\n".join(lines).strip()
        if len(text) > self.max_text_chars:
            raise ApiError(
                code="DOCUMENT_TEXT_LIMIT_EXCEEDED",
                message="Parsed document exceeds text processing limit.",
                status_code=422,
                details={"max_chars": self.max_text_chars},
            )

        warnings: list[str] = []
        if not text:
            warnings.append("xlsx_no_text_extracted")

        return ExtractionResult(
            text=text,
            page_count=sheet_count,
            extraction_method=self.extraction_method,
            coverage_score=1.0 if text else 0.0,
            warnings=warnings,
        )

    def _load_workbook_loader(self) -> Callable[..., object]:
        try:
            from openpyxl import load_workbook  # type: ignore[import-untyped]
        except Exception as exc:  # noqa: BLE001
            raise ApiError(
                code="XLSX_EXTRACTOR_DEPENDENCY_MISSING",
                message="openpyxl is required for XLSX extraction.",
                status_code=503,
            ) from exc
        return cast(Callable[..., object], load_workbook)
